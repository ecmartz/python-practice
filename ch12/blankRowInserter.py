#!python3

import openpyxl, string, sys

if __name__ == "__main__":
    start = sys.argv[1]
    blank = sys.argv[2]
    wb = openpyxl.load_workbook(sys.argv[3])
    sheet = wb.get_sheet_by_name('Sheet')

    alphabet = string.ascii_uppercase

    newwb = openpyxl.Workbook()
    newsheet = newwb.get_active_sheet()

    for i in range(0, sheet.max_column):
        for j in range(1, int(start)):
            cell = "%s%s" % (alphabet[i], j)
            newsheet[cell] = sheet[cell].value
        for j in range(int(start), int(start) + int(blank)):
            pass
        for j in range(int(start) + int(blank), sheet.max_row):
            cell = "%s%s" % (alphabet[i], j)
            newcell = "%s%s" % (alphabet[i], j)
            newsheet[newcell] = sheet[cell].value

    newwb.save("new-%s" % sys.argv[3])