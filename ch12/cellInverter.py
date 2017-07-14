#! python3
# cellInverter.py

import openpyxl, string, sys
import pdb

if __name__ == "__main__":
    file = sys.argv[1]
    alphabet = string.ascii_uppercase
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name('Sheet')

    newwb = openpyxl.Workbook()
    newsheet = newwb.get_active_sheet()

    for i in range(1, sheet.max_column + 1):
        for j in range(1, sheet.max_row + 1):
            cell = "%s%s" % (alphabet[i-1], j)
            newcell = "%s%s" % (alphabet[j-1], i)
            newsheet[newcell] = sheet[cell].value

    newwb.save("new-%s" % file)
